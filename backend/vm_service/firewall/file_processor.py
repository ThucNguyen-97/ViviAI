import hashlib
import re
import shutil
import uuid
import zipfile
from pathlib import Path
from typing import Sequence

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook, Workbook

from core.config import settings
from firewall.schemas import FirewallDecision, ProcessedFile, UserContext


FILE_REJECT_MESSAGE = "Tệp tin bạn gửi không phù hợp với chính sách hệ thống"
ALLOWED_EXTENSIONS = {".xlsx", ".png", ".md"}


def upload_raw_dir() -> Path:
    root = Path(settings.UPLOAD_DIR) / "raw"
    root.mkdir(parents=True, exist_ok=True)
    return root


def upload_clean_dir() -> Path:
    root = Path(settings.UPLOAD_DIR) / "clean"
    root.mkdir(parents=True, exist_ok=True)
    return root


def _max_size(extension: str) -> int:
    return {
        ".md": settings.MAX_MD_BYTES,
        ".png": settings.MAX_PNG_BYTES,
        ".xlsx": settings.MAX_XLSX_BYTES,
    }[extension]


def _safe_name(filename: str) -> str:
    name = Path(filename).name.strip()
    if not name or len(name) > 255 or any(ord(ch) < 32 for ch in name):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)
    return name


def _check_signature(data: bytes, raw_path: Path, extension: str) -> None:
    """Kiểm tra Signature Magic Bytes và cấu trúc tệp chuẩn bằng code Backend."""
    if extension == ".png":
        if not data.startswith(b"\x89PNG\r\n\x1a\n"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)
    elif extension == ".xlsx":
        if not data.startswith(b"PK\x03\x04") or not zipfile.is_zipfile(raw_path):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)
    elif extension == ".md":
        try:
            data.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)


def sanitize_md_for_prompt(md_content: str) -> str:
    """Xóa đường dẫn URL Markdown [text](url) -> text bằng Regex."""
    return re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", md_content)


def _sanitize_markdown(raw_path: Path, target_path: Path) -> None:
    text = raw_path.read_text(encoding="utf-8", errors="replace")
    clean_text = sanitize_md_for_prompt(text)
    target_path.write_text(clean_text, encoding="utf-8")


def _sanitize_xlsx(raw_path: Path, target_path: Path) -> None:
    workbook = load_workbook(
        raw_path,
        read_only=False,
        keep_vba=False,
        data_only=True,
        keep_links=False,
    )
    clean = Workbook()
    default = clean.active
    clean.remove(default)

    for source in workbook.worksheets:
        target = clean.create_sheet(title=source.title[:31] or "Sheet")
        for row in source.iter_rows():
            for cell in row:
                target[cell.coordinate].value = cell.value

    clean.save(target_path)


def _resize_and_strip_png(raw_path: Path, target_path: Path, max_dimension: int = 1024) -> bool:
    """Xóa sạch Metadata EXIF cho ảnh PNG và resize nếu quá lớn để tiết kiệm token."""
    try:
        from PIL import Image
        with Image.open(raw_path) as img:
            img_format = img.format or "PNG"
            img = img.convert("RGBA" if img.mode == "RGBA" else "RGB")
            
            w, h = img.size
            if w > max_dimension or h > max_dimension:
                if w > h:
                    new_w = max_dimension
                    new_h = int(h * (max_dimension / w))
                else:
                    new_h = max_dimension
                    new_w = int(w * (max_dimension / h))
                img = img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            
            img.save(target_path, format=img_format)
            return True
    except Exception:
        shutil.copyfile(raw_path, target_path)
        return False


async def process_uploads(files: Sequence[UploadFile], user: UserContext) -> list[ProcessedFile]:
    if len(files) > settings.MAX_FILES_PER_REQUEST:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)

    processed_files: list[ProcessedFile] = []
    for upload in files:
        original_name = _safe_name(upload.filename or "")
        extension = Path(original_name).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)

        data = await upload.read()
        if len(data) > _max_size(extension):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE)

        file_uuid = str(uuid.uuid4())
        raw_path = upload_raw_dir() / f"{file_uuid}{extension}"
        raw_path.write_bytes(data)

        # 1. Check Signature Magic Bytes & ZIP Structure
        _check_signature(data, raw_path, extension)

        clean_path = upload_clean_dir() / f"{file_uuid}{extension}"
        sanitized = True

        try:
            if extension == ".xlsx":
                _sanitize_xlsx(raw_path, clean_path)
            elif extension == ".md":
                _sanitize_markdown(raw_path, clean_path)
            elif extension == ".png":
                sanitized = _resize_and_strip_png(raw_path, clean_path, max_dimension=1024)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=FILE_REJECT_MESSAGE) from exc

        processed = ProcessedFile(
            original_file_name=original_name,
            file_type=extension.lstrip("."),
            raw_path=str(raw_path),
            clean_path=str(clean_path),
            mime_type=upload.content_type,
            sanitized=sanitized,
            flags=[],
            ek_file_id=file_uuid,
        )
        processed_files.append(processed)

    return processed_files

